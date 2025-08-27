import zipfile
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import io
from io import BytesIO
from datetime import datetime, date, timedelta
import math

def process_dataframe(df):
    try:
        # Fix the column mapping. CSV columns are shifted
        df_corrected = pd.DataFrame({
            'Gender': df['Load#'].str.strip(),        # Actual gender is in Load# column
            'Division': df['Gender'].str.strip(),     # Division is in Gender column  
            'Style': df['Division'].str.strip(),      # Style is in Division column
            'Color': df['Style'].str.strip(),         # Color is in Style column
            'Size': df['Color'].str.strip(),          # Size is in Color column
            'Qty': df[' "EU Size"']                   # Quantity is in "EU Size" column
        })
       
        # Clean up
        df_corrected['Gender'] = df_corrected['Gender'].str.strip()
        df_corrected['Style'] = df_corrected['Style'].astype(str).str.strip()
        df_corrected['Color'] = df_corrected['Color'].str.strip()
       
        # Check and make sure Qty is numeric
        df_corrected['Qty'] = pd.to_numeric(df_corrected['Qty'], errors='coerce') #If Qty value is not a number, turn to NaN
        df_corrected = df_corrected.dropna(subset=['Qty']) #Drop all NaN and only keep numbers
       
        # Group and sum
        # Groups data with same Gender, Style, Color and sums the Qty of each group
        grouped_df = df_corrected.groupby(['Gender', 'Style', 'Color'])['Qty'].sum().reset_index()
        # Changes Qty for Total and shows the total of each group 
        grouped_df.columns = ['Gender', 'Style', 'Color', 'Total']
       
        # Filter into categories
        men_df = grouped_df[grouped_df['Gender'] == 'MENS']
        women_df = grouped_df[grouped_df['Gender'] == 'WOMENS']
        kids_df = grouped_df[grouped_df['Gender'].isin(['BOYS', 'GIRLS']) & ~grouped_df['Style'].str.endswith('N')]
        work_df = grouped_df[grouped_df['Gender'] == 'INDUSTRIAL']
        nano_df = grouped_df[grouped_df['Style'].str.endswith('N')]
        
        # Calculate total quantity from original corrected dataframe
        total_qty = df_corrected['Qty'].sum()
        
        return men_df, women_df, kids_df, work_df, nano_df, total_qty, None
        
    except Exception as e:
        return None, None, None, None, None, 0, str(e)

def create_excel_file(men_df, women_df, kids_df, work_df, nano_df):
    # Create Excel file in memory and return as bytes bc idk
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        men_df.to_excel(writer, sheet_name='Men', index=False)
        women_df.to_excel(writer, sheet_name='Women', index=False)
        kids_df.to_excel(writer, sheet_name='Kids', index=False)
        nano_df.to_excel(writer, sheet_name='Nano', index=False)
        work_df.to_excel(writer, sheet_name='Work', index=False)
    
    return output.getvalue()

# Reservatron stuff
def style_cell(cell):
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(size=14, bold=True)

def style_ref(cell):
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(size=14, italic=True)    

def fill_customer_data(ws, customer, position):
    """Fill customer data in specific position on worksheet"""
    # Define cell coordinates for each position (1-10)
    positions = {
        1: {"name": "D4", "number": "E5", "obs": "C6", "worker": "D8", "date": "G8"},
        2: {"name": "N4", "number": "O5", "obs": "M6", "worker": "N8", "date": "Q8"},
        3: {"name": "D13", "number": "E14", "obs": "C15", "worker": "D17", "date": "G17"},
        4: {"name": "N13", "number": "O14", "obs": "M15", "worker": "N17", "date": "Q17"},
        5: {"name": "D22", "number": "E23", "obs": "C24", "worker": "D26", "date": "G26"},
        6: {"name": "N22", "number": "O23", "obs": "M24", "worker": "N26", "date": "Q26"},
        7: {"name": "D31", "number": "E32", "obs": "C33", "worker": "D35", "date": "G35"},
        8: {"name": "N31", "number": "O32", "obs": "M33", "worker": "N35", "date": "Q35"},
        9: {"name": "D40", "number": "E41", "obs": "C42", "worker": "D44", "date": "G44"},
        10: {"name": "N40", "number": "O41", "obs": "M42", "worker": "N44", "date": "Q44"}
    }
    
    coords = positions[position]
    
    # Fill the data and style it
    ws[coords["name"]] = f"{customer['NAME']} {customer['LAST NAME'] if pd.notna(customer['LAST NAME']) else ''}"
    style_cell(ws[coords["name"]])
    ws[coords["number"]] = customer["NUMBER"]
    style_cell(ws[coords["number"]])
    ws[coords["obs"]] = f"{customer["REF"]} {customer["COLOR"]} {customer["SIZE"]}"
    style_ref(ws[coords["obs"]])
    ws[coords["worker"]] = customer["WORKER"]
    style_cell(ws[coords["worker"]])
    ws[coords["date"]] = f"Fecha: {customer['DATE']}"
    style_ref(ws[coords["date"]])

def process_all_customers(client_df, template_file):
    total_customers = len(client_df)
    customers_per_page = 10
    num_files = math.ceil(total_customers / customers_per_page)

    outputs = []
    
    for file_num in range(num_files):
        # Load fresh template for each file
        wb = load_workbook(template_file)
        ws = wb['reserva_template']
        
        # Calculate customer range for this file
        start_idx = file_num * customers_per_page
        end_idx = min(start_idx + customers_per_page, total_customers)
        
        # Fill customers for this file
        for i in range(start_idx, end_idx):
            customer = client_df.iloc[i]
            position = (i - start_idx) + 1  # Position within this page (1-10)
            fill_customer_data(ws, customer, position)

        # Save each filled workbook into memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        outputs.append(output)
    
    return outputs

        

# Streamlit App
st.title('🚀 LISTATRON 505')

# Sidebar - Day Counter Utility
st.sidebar.title("🗓️ Contador de días")

# Start date input
start_date = st.sidebar.date_input("Fecha de compra")

# Calculate date 45 days from start_date
date_plus_45 = start_date + timedelta(days=45)

# End date input (optional)
use_today = st.sidebar.checkbox("Hasta HOY", value=True)

if use_today:
    end_date = date.today()
else:
    end_date = st.sidebar.date_input("Elegir fecha", min_value=start_date)

# Calculate difference and display
if start_date and end_date:
    delta = end_date - start_date
    
    # Show action dependent on time
    if delta.days > 45:
        st.sidebar.subheader(f"**{start_date.strftime('%d %b %Y')}** - **{end_date.strftime('%d %b %Y')}**")
        st.sidebar.title(f"📅 **{delta.days}** días")
        st.sidebar.markdown("---")
        st.sidebar.title(":red[❌ _No se puede hacer el cambio_]")
    elif delta.days == 1:
        st.sidebar.subheader(f"**{start_date.strftime('%d %b %Y')}** - **{end_date.strftime('%d %b %Y')}**")
        st.sidebar.title(f"📅 **{delta.days}** día")
        st.sidebar.markdown("---")
        st.sidebar.title(":green[✅ _Se puede hacer el cambio_]")
    elif delta.days > 1 and delta.days <= 45:
        st.sidebar.subheader(f"**{start_date.strftime('%d %b %Y')}** - **{end_date.strftime('%d %b %Y')}**")
        st.sidebar.title(f"📅 **{delta.days}** días")
        st.sidebar.markdown("---")
        st.sidebar.title(":green[✅ _Se puede hacer el cambio_]")
    elif delta.days < 0:
        st.sidebar.title(":red[Error!❌ Esa fecha es en el futuro]")
        st.sidebar.subheader(f"Faltan {abs(delta.days)} dias para llegar")
    elif delta.days == 0:
        st.sidebar.markdown("---")
        st.sidebar.title(f"Es hoy!")
        st.sidebar.markdown(f"📅 **45 días después**: {date_plus_45.strftime('%d/%m/%Y')}")

tab1, tab2 = st.tabs(["Listatron", "Reservatron"])

with tab1:

    st.write('Carga el archivo y se descarga ordenado papá')

    # File uploader
    uploaded_file = st.file_uploader("Elegí el archivo", type="csv")

    if uploaded_file is not None:
        try:
            # Read the CSV file
            df = pd.read_csv(uploaded_file)
            
            # Process the data
            men_df, women_df, kids_df, work_df, nano_df, total_qty, error = process_dataframe(df)
            
            if error:
                st.error(f"Error: {error}")
            else:
                st.success(f"Se subio tuani! Total pares: {int(total_qty)}")

                # Create and offer download
                excel_data = create_excel_file(men_df, women_df, kids_df, work_df, nano_df)
                
                st.markdown("---")
                st.subheader("📥 Descargame dog")
                st.download_button(
                    label="📥 Descargar archivo ordenado",
                    data=excel_data,
                    file_name="data_sorted.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.success("Se guardó tuani! Descargalo dog")
                st.markdown("---")

                st.subheader("Preview para revisar que todo Gucci")
                
                # Show summary with total Qty for each category
                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    men_total = men_df['Total'].sum() if not men_df.empty else 0
                    st.metric("Men", f"{int(men_total)}", f"{len(men_df)} refes")
                with col2:
                    women_total = women_df['Total'].sum() if not women_df.empty else 0
                    st.metric("Women", f"{int(women_total)}", f"{len(women_df)} refes")
                with col3:
                    kids_total = kids_df['Total'].sum() if not kids_df.empty else 0
                    st.metric("Kids", f"{int(kids_total)}", f"{len(kids_df)} refes")
                with col4:
                    work_total = work_df['Total'].sum() if not work_df.empty else 0
                    st.metric("Work", f"{int(work_total)}", f"{len(work_df)} refes")
                with col5:
                    nano_total = nano_df['Total'].sum() if not nano_df.empty else 0
                    st.metric("Nano", f"{int(nano_total)}", f"{len(nano_df)} refes")
                
                # Show preview of each category
                categories = {
                    "Men": men_df,
                    "Women": women_df,
                    "Kids": kids_df,
                    "Work": work_df,
                    "Nano": nano_df
                }
                
                for category, data in categories.items():
                    if not data.empty:
                        with st.expander(f"Preview {category} data"):
                            st.dataframe(data)
                

        except Exception as e:
            st.error(f"Error leyendo: {str(e)}")
    else:
        st.info("Tiene que ser CSV")

with tab2:
    st.write("A llenar reservas")

    uploaded_data = st.file_uploader("Dame la info de reservas (.xlsx)", type=["xlsx"])
    uploaded_template = st.file_uploader("Dame el template (.xlsx)", type=["xlsx"])

    if uploaded_data and uploaded_template:
        client_df = pd.read_excel(uploaded_data)

        if st.button("Generate forms"):
            outputs = process_all_customers(client_df, uploaded_template)

            # If only 1 file show one download button
            if len(outputs) == 1:
                st.download_button("Download forms", data=outputs[0], file_name="filled.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet") 
            else:
                # Multiple files offer multiple download buttons or zip it
                zip_buffer = BytesIO()
                with zipfile.ZipFile(zip_buffer, "w") as zf:
                    for idx, out in enumerate(outputs, start=1):
                        zf.writestr(
                            f"filled_{idx}.xlsx",
                            out.getvalue()
                        )
                zip_buffer.seek(0)
                st.markdown("Esto descarga un zip con todos los archivos")
                st.download_button(
                    "Download ALL",
                    data=zip_buffer,
                    file_name="all_forms.zip",
                    mime="application/zip"
                )

                # Or download singles
                st.markdown("Y aqui separados")
                for idx, out in enumerate(outputs, start=1):
                    st.download_button(f"Download Page {idx}", data=out, file_name=f"filled_from_{idx}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
st.markdown("---")
st.markdown("*SM type shit*")